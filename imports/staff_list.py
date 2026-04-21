"""
imports/staff_list.py
Imports staff from the staff list Excel file into the database.

Expected columns (case-insensitive):
    Horizon Person Number
    Name
    Job Title
    Job Family
    Job Function
    Department
    Availability
    Start Date
    End Date

Job Function is the discipline — derived from the suffix after the comma
in the Horizon job title e.g. "Lead Professional, Mechanical Engineering"
becomes "Mechanical Engineering". The column can also be populated manually
for staff whose job title doesn't follow this pattern.

Run directly:
    python imports/staff_list.py source-data/staff_list.xlsx
"""

import json
import openpyxl
from datetime import datetime, timezone, date
import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database import get_connection
import config


# ---------------------------------------------------------------------------
# Column name normalisation
# ---------------------------------------------------------------------------
COLUMN_MAP = {
    "horizon person number": "horizon_person_number",
    "name":                  "name",
    "job title":             "job_title",
    "job family":            "job_family",
    "job function":          "job_function",
    "department":            "department",
    "availability":          "availability",
    "start date":            "start_date",
    "end date":              "end_date",
}


def _normalise_header(h):
    if h is None:
        return ""
    return str(h).strip().lower()


def _clean(val):
    if val is None:
        return None
    if isinstance(val, str):
        v = val.strip()
        return v if v else None
    return val


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        if hasattr(val, "date"):
            return val.date().isoformat()
        return val.isoformat()
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _parse_availability(val):
    if val is None:
        return 1.0
    try:
        f = float(val)
        return max(0.0, min(1.0, f))
    except (ValueError, TypeError):
        return 1.0


def _job_function_from_title(job_title):
    """
    Extract job function from job title suffix.
    "Lead Professional, Mechanical Engineering" -> "Mechanical Engineering"
    Returns None if no comma found.
    """
    if not job_title:
        return None
    if "," in job_title:
        return job_title.split(",", 1)[1].strip() or None
    return None


def run(file_path: str) -> dict:
    started_at = datetime.now(timezone.utc).isoformat()
    errors     = []
    processed  = 0
    inserted   = 0
    updated    = 0
    skipped    = 0

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        return _log(file_path, started_at, 0, 0, 0, 0, [str(e)])

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return _log(file_path, started_at, 0, 0, 0, 0, ["Sheet has no data rows"])

    # Build column index from header row
    raw_headers = [_normalise_header(h) for h in rows[0]]
    col_idx = {}
    for i, h in enumerate(raw_headers):
        field = COLUMN_MAP.get(h)
        if field:
            col_idx[field] = i

    # Horizon person number is required
    if "horizon_person_number" not in col_idx:
        return _log(file_path, started_at, 0, 0, 0, 0,
                    ["Column 'Horizon Person Number' not found in header row"])

    def get(row, field):
        idx = col_idx.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    conn = get_connection()
    c    = conn.cursor()
    now  = datetime.now(timezone.utc).isoformat()

    for row in rows[1:]:
        if not any(row):
            continue

        processed += 1

        horizon_id = _clean(get(row, "horizon_person_number"))
        name       = _clean(get(row, "name"))

        # Skip placeholder rows (no horizon ID or no name)
        if not horizon_id or not name:
            skipped += 1
            continue

        job_title    = _clean(get(row, "job_title"))
        job_family   = _clean(get(row, "job_family"))
        job_function = _clean(get(row, "job_function"))
        department   = _clean(get(row, "department"))
        availability = _parse_availability(get(row, "availability"))
        start_date   = _parse_date(get(row, "start_date"))
        end_date     = _parse_date(get(row, "end_date"))

        # If job_function not explicitly provided, try to derive from job title
        if not job_function and job_title:
            job_function = _job_function_from_title(job_title)

        existing = c.execute(
            "SELECT horizon_person_number FROM staff WHERE horizon_person_number = ?",
            (horizon_id,)
        ).fetchone()

        if existing:
            c.execute("""
                UPDATE staff SET
                    name=?, job_title=?, job_family=?, job_function=?,
                    department=?, availability=?, start_date=?,
                    end_date=?, last_imported=?
                WHERE horizon_person_number=?
            """, (name, job_title, job_family, job_function,
                  department, availability, start_date,
                  end_date, now, horizon_id))
            updated += 1
        else:
            c.execute("""
                INSERT INTO staff (
                    horizon_person_number, name, job_title, job_family,
                    job_function, department, availability,
                    start_date, end_date, last_imported
                ) VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (horizon_id, name, job_title, job_family,
                  job_function, department, availability,
                  start_date, end_date, now))
            inserted += 1

    conn.commit()
    conn.close()

    return _log(file_path, started_at, processed, inserted, updated,
                skipped, errors)


def _log(file_path, started_at, processed, inserted, updated, skipped, errors):
    completed_at = datetime.now(timezone.utc).isoformat()
    result = {
        "import_type":    "staff_list",
        "filename":       os.path.basename(str(file_path)),
        "started_at":     started_at,
        "completed_at":   completed_at,
        "rows_processed": processed,
        "rows_inserted":  inserted,
        "rows_updated":   updated,
        "rows_skipped":   skipped,
        "errors":         errors,
    }
    try:
        conn = get_connection()
        conn.execute("""
            INSERT INTO import_log
                (import_type, filename, started_at, completed_at,
                 rows_processed, rows_inserted, rows_updated, errors)
            VALUES (?,?,?,?,?,?,?,?)
        """, (result["import_type"], result["filename"],
              result["started_at"], result["completed_at"],
              result["rows_processed"], result["rows_inserted"],
              result["rows_updated"], json.dumps(result["errors"])))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Warning: could not write import log: {e}")
    return result


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else str(config.STAFF_LIST_PATH)
    result = run(path)
    print(f"Staff import:")
    print(f"  Source    : {result['filename']}")
    print(f"  Processed : {result['rows_processed']}")
    print(f"  Inserted  : {result['rows_inserted']}")
    print(f"  Updated   : {result['rows_updated']}")
    print(f"  Skipped   : {result['rows_skipped']} (no Horizon ID or name)")
    if result["errors"]:
        print(f"  Errors ({len(result['errors'])}):")
        for e in result["errors"][:5]:
            print(f"    - {e}")
