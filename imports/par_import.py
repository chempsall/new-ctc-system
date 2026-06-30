"""
imports/par_import.py
Imports the Project Attribute Report (PAR) from Horizon into the projects table.

The PAR is the authoritative source for all project metadata and financial
figures. This import runs nightly and populates/updates the projects table.
The CTC macro push then adds resource allocation data against those projects.

TWO MODES — controlled by RF_PAR_USE_SHAREPOINT in .env:

  RF_PAR_USE_SHAREPOINT=false (default/testing):
    Reads the most recent UK_PAR*.xlsx file from the source-data folder.
    Filename must start with YYYYMM e.g. 202604_UK_PAR_Active.xlsx

  RF_PAR_USE_SHAREPOINT=true (production):
    Connects to SharePoint via Microsoft Graph API.
    Finds the latest UK_PAR file automatically.
    Requires RF_SP_TENANT_ID, RF_SP_CLIENT_ID, RF_SP_CLIENT_SECRET in .env.

Filtering: Project Status = "Active" only.
Three indirect rows always added (annual leave, training, study leave).
"""

import json
import io
import openpyxl
import re
from datetime import datetime, timezone
import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database import get_connection
import config


# ---------------------------------------------------------------------------
# Column mapping: our field name -> PAR column header
# ---------------------------------------------------------------------------
PAR_COLUMNS = {
    "project_type":             "Project Type",
    "project_number":           "Project Number",
    "project_name":             "Project Name",
    "project_organisation":     "Project Organization",
    "project_customer":         "Project Customer",
    "project_status":           "Project Status",
    "project_director":         "Project Director",
    "project_manager":          "Project Manager",
    "task_number":              "Task Number",
    "task_name":                "Task Name",
    "task_start_date":          "Task Start Date",
    "task_end_date":            "Task End Date",
    "reporting_period":         "Reporting Period",
}

# Three indirect rows always added — replicates the original M code ManualRows
INDIRECT_ROWS = [
    {
        "project_type": "UK Indirect", "project_number": "IDUK-01",
        "project_name": "Learning Day Release & Study Leave",
        "task_number": "IDUK-01", "task_name": "Learning Day Release & Study Leave",
        "project_status": "Active",
        "project_organisation": "UK010117-UK-BSV-Services London",
        "project_customer": "UK010117-UK-BSV-Services London",
    },
    {
        "project_type": "UK Indirect", "project_number": "ID-04",
        "project_name": "Training Received",
        "task_number": "ID-04", "task_name": "Training Received",
        "project_status": "Active",
        "project_organisation": "UK010117-UK-BSV-Services London",
        "project_customer": "UK010117-UK-BSV-Services London",
    },
    {
        "project_type": "UK Indirect", "project_number": "ID-06",
        "project_name": "Annual Leave & Public Holiday",
        "task_number": "ID-06", "task_name": "Annual Leave & Public Holiday",
        "project_status": "Active",
        "project_organisation": "UK010117-UK-BSV-Services London",
        "project_customer": "UK010117-UK-BSV-Services London",
    },
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _prefix_from_filename(filename):
    m = re.match(r'^(\d{6})', os.path.basename(filename))
    return m.group(1) if m else None


def _clean(val):
    if val is None:
        return None
    if isinstance(val, str):
        v = val.strip().lstrip("'").strip()
        return v if v else None
    return val


def _clean_name(val):
    """Strip parenthetical suffixes: 'Smith, John (John)' -> 'Smith, John'"""
    v = _clean(val)
    return v.split("(")[0].strip() if v else None


def _parse_date(val):
    if val is None:
        return None
    if hasattr(val, "year"):
        return val.date().isoformat() if hasattr(val, "date") else val.isoformat()
    s = str(val).strip()
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s if s else None


# ---------------------------------------------------------------------------
# File acquisition
# ---------------------------------------------------------------------------

def _get_file_local():
    """
    Find the most recent UK_PAR*.xlsx in the configured directory.
    Falls back to using the configured path directly if it's a file.
    """
    configured = Path(config.PAR_ACTUALS_PATH)
    if configured.is_file():
        return str(configured), _prefix_from_filename(configured.name)

    search_dir = configured if configured.is_dir() else configured.parent
    candidates = []
    for f in search_dir.glob("*.xlsx"):
        prefix = _prefix_from_filename(f.name)
        if prefix and "UK_PAR" in f.name.upper():
            candidates.append((prefix, f))

    if not candidates:
        raise FileNotFoundError(
            f"No UK_PAR*.xlsx files found in {search_dir}.\n"
            f"Download the latest PAR file from SharePoint and place it there.\n"
            f"Filename must start with YYYYMM e.g. 202604_UK_PAR_Active.xlsx"
        )
    candidates.sort(key=lambda x: x[0], reverse=True)
    best_prefix, best_path = candidates[0]
    return str(best_path), best_prefix


def _get_file_sharepoint():
    """Connect to SharePoint and return the latest UK_PAR file as BytesIO."""
    try:
        import msal
        import requests
    except ImportError:
        raise ImportError(
            "msal and requests required for SharePoint. "
            "Install: pip install msal requests"
        )
    if not all([config.SHAREPOINT_TENANT_ID,
                config.SHAREPOINT_CLIENT_ID,
                config.SHAREPOINT_CLIENT_SECRET]):
        raise ValueError(
            "SharePoint credentials not configured. "
            "Set RF_SP_TENANT_ID, RF_SP_CLIENT_ID, RF_SP_CLIENT_SECRET in .env"
        )
    authority = f"https://login.microsoftonline.com/{config.SHAREPOINT_TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        config.SHAREPOINT_CLIENT_ID, authority=authority,
        client_credential=config.SHAREPOINT_CLIENT_SECRET,
    )
    token = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )
    if "access_token" not in token:
        raise ValueError(f"SharePoint auth failed: {token.get('error_description')}")

    hdrs     = {"Authorization": f"Bearer {token['access_token']}"}
    site_url = config.PAR_SHAREPOINT_SITE
    hostname = site_url.split("/")[2]
    sp_path  = "/".join(site_url.split("/")[3:])

    import requests as req
    site_id  = req.get(
        f"https://graph.microsoft.com/v1.0/sites/{hostname}:/{sp_path}",
        headers=hdrs).json()["id"]
    drive_id = req.get(
        f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives",
        headers=hdrs).json()["value"][0]["id"]

    folder = config.PAR_SHAREPOINT_FOLDER.strip("/")
    files  = req.get(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{folder}:/children",
        headers=hdrs).json()["value"]

    candidates = [
        (_prefix_from_filename(f["name"]), f)
        for f in files
        if "UK_PAR" in f.get("name","").upper() and _prefix_from_filename(f["name"])
    ]
    if not candidates:
        raise FileNotFoundError(f"No UK_PAR files in SharePoint folder: {folder}")

    candidates.sort(key=lambda x: x[0], reverse=True)
    best_prefix, best_file = candidates[0]
    print(f"  SharePoint: using {best_file['name']}")
    content = req.get(best_file["@microsoft.graph.downloadUrl"]).content
    return io.BytesIO(content), best_prefix


# ---------------------------------------------------------------------------
# Workbook parsing
# ---------------------------------------------------------------------------

def _parse_workbook(wb_source):
    """
    Parse PAR workbook. Returns list of dicts, filtered to Active projects only.
    """
    if isinstance(wb_source, (str, Path)):
        wb = openpyxl.load_workbook(str(wb_source), read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(wb_source, read_only=True, data_only=True)

    ws   = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header  = [str(h).strip() if h else "" for h in rows[0]]
    col_idx = {h: i for i, h in enumerate(header)}

    # Build field->column_index map
    field_map = {}
    for field, par_col in PAR_COLUMNS.items():
        idx = col_idx.get(par_col)
        if idx is not None:
            field_map[field] = idx

    status_idx = col_idx.get("Project Status")

    data = []
    for row in rows[1:]:
        if not any(row):
            continue
        # Filter to Active only (replicates M code FilteredRows step)
        if status_idx is not None:
            status = _clean(row[status_idx])
            if status and status.lower() != "active":
                continue
        d = {field: row[idx] for field, idx in field_map.items()}
        data.append(d)

    return data


# ---------------------------------------------------------------------------
# Database upsert
# ---------------------------------------------------------------------------

def _upsert_projects(data_rows, now):
    """Upsert all rows into projects table. Returns (inserted, updated, errors)."""
    conn     = get_connection()
    c        = conn.cursor()
    inserted = updated = 0
    errors   = []

    for d in data_rows + INDIRECT_ROWS:
        proj_num = _clean(d.get("project_number"))
        task_num = _clean(d.get("task_number"))

        if not proj_num or not task_num:
            errors.append("Skipped: missing project_number or task_number")
            continue

        vals = (
            _clean(d.get("project_type")),
            _clean(d.get("project_name")),
            _clean(d.get("task_name")),
            _clean(d.get("project_organisation")),
            _clean(d.get("project_customer")),
            _clean(d.get("project_status")) or "Active",
            _clean_name(d.get("project_director")),
            _clean_name(d.get("project_manager")),
            _parse_date(d.get("task_start_date")),
            _parse_date(d.get("task_end_date")),
            _clean(d.get("reporting_period")),
            now,
        )

        existing = c.execute("""
            SELECT project_id FROM projects
            WHERE project_number = ? AND task_order_number = ?
        """, (proj_num, task_num)).fetchone()

        if existing:
            c.execute("""
                UPDATE projects SET
                    project_type=?, project_name=?, task_name=?,
                    project_organisation=?, project_customer=?,
                    project_status=?, project_director=?, project_manager=?,
                    task_start_date=?, task_end_date=?, reporting_period=?,
                    last_imported=?
                WHERE project_id=?
            """, vals + (existing["project_id"],))
            updated += 1
        else:
            c.execute("""
                INSERT INTO projects (
                    project_number, task_order_number,
                    project_type, project_name, task_name,
                    project_organisation, project_customer, project_status,
                    project_director, project_manager,
                    task_start_date, task_end_date, reporting_period,
                    last_imported
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (proj_num, task_num) + vals)
            inserted += 1

    conn.commit()
    conn.close()
    return inserted, updated, errors


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def run(file_path=None):
    """
    Import PAR data into the projects table.
    Mode (local file vs SharePoint) is controlled by RF_PAR_USE_SHAREPOINT in .env.
    """
    started_at   = datetime.now(timezone.utc).isoformat()
    source_label = ""
    errors       = []

    use_sharepoint = os.environ.get(
        "RF_PAR_USE_SHAREPOINT", "false"
    ).lower() == "true"

    try:
        if use_sharepoint:
            print("  PAR import: connecting to SharePoint...")
            wb_source, _prefix = _get_file_sharepoint()
            source_label = "SharePoint"
        else:
            local_path, _prefix = _get_file_local()
            wb_source    = local_path
            source_label = os.path.basename(local_path)
            print(f"  PAR import: reading {source_label}")

        data_rows = _parse_workbook(wb_source)

        if not data_rows:
            return _log(source_label, started_at, 0, 0, 0,
                        ["No Active project rows found"])

        now = datetime.now(timezone.utc).isoformat()
        inserted, updated, row_errors = _upsert_projects(data_rows, now)
        errors.extend(row_errors)
        processed = len(data_rows) + len(INDIRECT_ROWS)

    except Exception as e:
        return _log(source_label or "unknown", started_at, 0, 0, 0, [str(e)])

    return _log(source_label, started_at, processed, inserted, updated, errors)


def _log(filename, started_at, processed, inserted, updated, errors):
    completed_at = datetime.now(timezone.utc).isoformat()
    result = {
        "import_type":    "par_import",
        "filename":       filename,
        "started_at":     started_at,
        "completed_at":   completed_at,
        "rows_processed": processed,
        "rows_inserted":  inserted,
        "rows_updated":   updated,
        "errors":         errors,
    }
    try:
        conn = get_connection()
        conn.execute("""
            INSERT INTO import_log
                (import_type, filename, started_at, completed_at,
                 rows_processed, rows_inserted, rows_updated, errors)
            VALUES (?,?,?,?,?,?,?,?)
        """, (result["import_type"], result["filename"],
              result["started_at"], result["completed_at"],
              result["rows_processed"], result["rows_inserted"],
              result["rows_updated"], json.dumps(result["errors"])))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Warning: could not write import log: {e}")
    return result


if __name__ == "__main__":
    result = run()
    print(f"PAR import:")
    print(f"  Source    : {result['filename']}")
    print(f"  Processed : {result['rows_processed']}")
    print(f"  Inserted  : {result['rows_inserted']}")
    print(f"  Updated   : {result['rows_updated']}")
    if result["errors"]:
        print(f"  Errors ({len(result['errors'])}):")
        for e in result["errors"][:5]:
            print(f"    - {e}")