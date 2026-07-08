"""
CTC Migration Tool
==================
Reads existing CTC Excel files and imports them into the RFT 3.0 database.

Usage:
    python ctc_migrate.py

Configuration (edit the constants below before running):
    CTC_FOLDER   — folder containing the CTC Excel files
    DB_PATH      — path to your RFT 3.0 database
    FROM_PERIOD  — only import allocations from this month onwards (YYYY-MM-01)
    DRY_RUN      — set True to preview without writing to the database
"""

import os
import sys
import re
import sqlite3
from pathlib import Path
from datetime import date, datetime, timezone

# ── Configuration ────────────────────────────────────────────────────────────

CTC_FOLDER  = r"C:\CTC-files"
DB_PATH     = r"C:\Users\UKCMH001\Dev\new-ctc-system\data\resource_forecast.db"
FROM_PERIOD  = "2026-01-01"   # import allocations from this date (includes history)
FUTURE_CUTOFF = "2026-07-01"  # files must have work from this date to be imported
DRY_RUN     = True           # set False to actually write to the database
DEFAULT_DEPT = "UK010117-UK-BSV-Services London"  # used when Horizon lookup fails

# ── Dependencies check ───────────────────────────────────────────────────────

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Run: pip install openpyxl")
    sys.exit(1)

# ── Helpers ──────────────────────────────────────────────────────────────────

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    return conn


def parse_period_label(label):
    """Convert 'Jul-26', date, or datetime -> '2026-07-01', returns None if unrecognised."""
    if not label:
        return None
    # Handle Excel date objects returned by openpyxl
    if hasattr(label, 'year') and hasattr(label, 'month'):
        return f"{label.year:04d}-{label.month:02d}-01"
    s = str(label).strip()
    # Match Mon-YY or Mon-YYYY
    m = re.match(r'^([A-Za-z]{3})-(\d{2,4})$', s)
    if not m:
        return None
    month_abbr = m.group(1).capitalize()
    year_str   = m.group(2)
    year = int(year_str) + 2000 if len(year_str) == 2 else int(year_str)
    months = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
              "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
    month = months.get(month_abbr)
    if not month:
        return None
    return f"{year:04d}-{month:02d}-01"


def cell_val(ws, cell_ref):
    """Get string value from a cell, stripping whitespace."""
    val = ws[cell_ref].value
    if val is None:
        return ""
    return str(val).strip()


def clean_name(raw):
    """Normalise a staff name — remove parenthetical duplicates like 'Smith, John (John)'."""
    if not raw:
        return ""
    # Remove trailing (firstname) pattern
    cleaned = re.sub(r'\s*\([^)]+\)\s*$', '', raw).strip()
    return cleaned


def is_placeholder(s):
    if not s:
        return True
    c = s.lower().strip()
    return c in {"xxxxxxxx", "00000000", "12345678", "tbc", "tbd", "n/a", ""}


def make_suffix():
    """Generate a timestamp suffix for placeholder IDs."""
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%f")[:20]

# ── Main processing ───────────────────────────────────────────────────────────

def process_file(path, conn, stats):
    """
    Process a single CTC Excel file.
    Returns a dict with the results for this file.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return {"file": path.name, "status": "ERROR", "error": str(e)}

    if "Resources" not in wb.sheetnames:
        return {"file": path.name, "status": "SKIPPED", "error": "No 'Resources' sheet"}

    ws = wb["Resources"]

    # Read header fields
    proj_num  = cell_val(ws, "G4")
    task_num  = cell_val(ws, "G5")
    dept      = cell_val(ws, "G7")
    if not dept or "No Horizon Record Found" in dept:
        dept = DEFAULT_DEPT
    pd_raw    = clean_name(cell_val(ws, "G11"))
    pm_raw    = clean_name(cell_val(ws, "G12"))

    if not proj_num:
        proj_num = "00000000"
    if not task_num:
        task_num = "000"

    # Read month headers from row 15, columns J onwards
    # Find which columns correspond to which periods
    period_cols = {}  # period_start -> column index (1-based)
    for col in range(10, 60):  # columns J (10) to BH (60)
        cell = ws.cell(row=15, column=col)
        period = parse_period_label(cell.value)
        if period and period >= FUTURE_CUTOFF:
            period_cols[period] = col

    # period_cols may be empty if all allocations are in the past — that's fine,
    # we'll still create the RTC with minimum 12 periods and zero allocations

    # Read staff rows F16:F55
    staff_allocations = []  # list of {name, allocations: {period: days}}
    for row in range(16, 56):
        name_cell = ws.cell(row=row, column=6)  # column F
        name = str(name_cell.value).strip() if name_cell.value else ""
        if not name:
            continue

        allocations = {}
        for period, col in period_cols.items():
            val = ws.cell(row=row, column=col).value
            try:
                days = float(val) if val is not None else 0.0
            except (ValueError, TypeError):
                days = 0.0
            if days > 0:
                allocations[period] = days

        if allocations:  # only include staff with actual allocations in scope
            staff_allocations.append({"name": name, "allocations": allocations})

    # Check if any staff have future allocations
    has_future_work = any(s["allocations"] for s in staff_allocations)
    
    if not has_future_work:
        return {"file": path.name, "status": "SKIPPED",
                "error": "No future allocations — excluded from migration"}

    # Calculate start and target periods from actual allocation data
    from datetime import date as _date
    from dateutil.relativedelta import relativedelta as _rd
    _periods_with_data = [
        period for s in staff_allocations
        for period in s["allocations"].keys()
    ]
    if _periods_with_data:
        _first_alloc = min(_periods_with_data)
        _last_alloc  = max(_periods_with_data)
    else:
        from datetime import date as _d2
        _current_month = _d2.today().replace(day=1).isoformat()
        _first_alloc = _current_month
        _last_alloc  = _current_month

    start_period = _first_alloc
    _start  = _date.fromisoformat(start_period)
    _end    = _date.fromisoformat(_last_alloc)
    target  = _start + _rd(months=11)
    _blocks = 1
    while target < _end:
        target  += _rd(months=12)
        _blocks += 1

    print(f"  DEBUG: start={start_period} end={_last_alloc} target={target} blocks={_blocks}")

    result = {
        "file":        path.name,
        "proj_num":    proj_num,
        "task_num":    task_num,
        "dept":        dept,
        "pd":          pd_raw,
        "pm":          pm_raw,
        "staff":       len(staff_allocations),
        "periods":     sorted(period_cols.keys()),
        "grid_months": _blocks * 12,
        "start_period": start_period,
    }

    if DRY_RUN:
        result["status"] = "DRY_RUN"
        result["allocations_preview"] = [
            {"name": s["name"], "total_days": sum(s["allocations"].values())}
            for s in staff_allocations
        ]
        return result

    # ── Write to database ────────────────────────────────────────────────────
    c   = conn.cursor()
    now = datetime.now(timezone.utc).isoformat()

    # Look up project in database
    # Try exact match first
    proj_row = c.execute("""
        SELECT project_id, project_name FROM projects
        WHERE project_number = ? AND task_order_number = ?
        AND project_status = 'Active'
    """, (proj_num, task_num)).fetchone()

    if not proj_row and not is_placeholder(task_num):
        # Try project-only match — only if task number is real
        # (placeholder task numbers should never share an existing project)
        proj_row = c.execute("""
            SELECT project_id, project_name FROM projects
            WHERE project_number = ?
            AND project_status = 'Active'
            LIMIT 1
        """, (proj_num,)).fetchone()

    if proj_row:
        project_id = proj_row["project_id"]
        result["horizon"] = "linked"
        result["project_name"] = proj_row["project_name"]
    else:
        # Create a Pending project row
        suffix = make_suffix()
        unique_task = f"{task_num}_{suffix}" if not is_placeholder(task_num) else f"PLACEHOLDER_{suffix}"
        c.execute("""
            INSERT INTO projects
                (project_number, task_order_number, project_name, task_name,
                 project_customer, project_director, project_manager,
                 project_status, last_imported)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'Pending', ?)
        """, (proj_num, unique_task,
              f"{proj_num} / {task_num}",
              "",
              None,
              None if not pd_raw or "No Horizon Record Found" in pd_raw else pd_raw,
              None if not pm_raw or "No Horizon Record Found" in pm_raw else pm_raw,
              now))
        project_id = c.lastrowid
        result["horizon"] = "pending"

    # Check if RTC already exists for this project — flag as collision
    existing_rtc = c.execute("""
        SELECT rtc_id FROM rtcs WHERE project_id = ?
    """, (project_id,)).fetchone()

    if existing_rtc:
        result["status"]    = "COLLISION"
        result["error"]     = f"RTC already exists for {proj_num}/{task_num} — possible duplicate project number"
        conn.rollback()
        return result

    if existing_rtc:
        rtc_id = existing_rtc["rtc_id"]
        result["rtc"] = "existing"
    else:
        # Create new RTC
        c.execute("""
            INSERT INTO rtcs
                (project_id, department, start_date,
                 created_by, created_at, last_updated_by, last_updated_at,
                 is_archived, auto_linked)
            VALUES (?, ?, ?, 'Migration', ?, 'Migration', ?, 0, 0)
        """, (project_id, dept, start_period, now, now))
        rtc_id = c.lastrowid
        result["rtc"] = "created"

    # Ensure reporting periods exist through target
    import database as db_module
    db_module.ensure_periods_through(conn, target)

    # Add staff and allocations
    rows_added = 0
    staff_skipped = 0
    for s in staff_allocations:
        # Look up staff member by name (last, first format)
        staff_row = c.execute("""
            SELECT horizon_person_number FROM staff
            WHERE name = ? AND (end_date IS NULL OR end_date > ?)
        """, (s["name"], FROM_PERIOD)).fetchone()

        if not staff_row:
            # Try partial name match
            staff_row = c.execute("""
                SELECT horizon_person_number FROM staff
                WHERE name LIKE ? AND (end_date IS NULL OR end_date > ?)
                LIMIT 1
            """, (f"%{s['name'].split(',')[0].strip()}%", FROM_PERIOD)).fetchone()

        if not staff_row:
            staff_skipped += 1
            stats["staff_skipped"].append(s["name"])
            continue

        pid = staff_row["horizon_person_number"]

        # Seed zero rows for all periods up to target
        all_periods = c.execute("""
            SELECT period_start FROM reporting_periods
            WHERE period_start >= ? AND period_start <= ?
            ORDER BY period_start
        """, (start_period, target.isoformat())).fetchall()

        for p_row in all_periods:
            period = p_row["period_start"]
            days = s["allocations"].get(period, 0)
            c.execute("""
                INSERT OR IGNORE INTO allocations
                    (horizon_person_number, rtc_id, period_start, days, last_updated)
                VALUES (?, ?, ?, ?, ?)
            """, (pid, rtc_id, period, days, now))
            if days > 0:
                c.execute("""
                    UPDATE allocations SET days = ?, last_updated = ?
                    WHERE horizon_person_number = ? AND rtc_id = ? AND period_start = ?
                """, (days, now, pid, rtc_id, period))
                rows_added += 1

    conn.commit()
    result["status"]        = "OK"
    result["rows_added"]    = rows_added
    result["staff_skipped"] = staff_skipped
    return result

def preflight_check(files):
    """Scan all files and report duplicate project+task combinations."""
    print("\nPre-flight check — scanning for duplicate project/task numbers...")
    seen = {}  # (proj_num, task_num) -> list of filenames
    for f in files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
            if "Resources" not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["Resources"]
            proj_num = str(ws["G4"].value or "").strip() or "00000000"
            task_num = str(ws["G5"].value or "").strip() or "000"
            wb.close()
            key = (proj_num, task_num)
            seen.setdefault(key, []).append(f.name)
        except Exception:
            pass

    duplicates = {k: v for k, v in seen.items() if len(v) > 1}
    if duplicates:
        print(f"\nWARNING: {len(duplicates)} duplicate project/task combinations found:")
        for (proj, task), fnames in sorted(duplicates.items()):
            print(f"  {proj} / {task}:")
            for fname in fnames:
                print(f"    - {fname}")
        print()
        return False
    else:
        print(f"  OK — no duplicates found across {len(files)} files\n")
        return True

def main():
    folder = Path(CTC_FOLDER)
    if not folder.exists():
        print(f"ERROR: Folder not found: {CTC_FOLDER}")
        sys.exit(1)

    all_files = sorted(folder.glob("*.xlsm"))
    if not all_files:
        print(f"ERROR: No .xlsm files found in {CTC_FOLDER}")
        sys.exit(1)

    # Step 1: Filter to files with future work only
    print(f"Scanning {len(all_files)} files for future allocations...")
    files = []
    no_future = []
    for f in all_files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
            if "Resources" not in wb.sheetnames:
                no_future.append(f.name)
            if has_future:
                if f.name == "UK0042369.4312-UK-DK01 Bulk BMS and ICT.xlsm":
                    print(f"  DEBUG: passed filter, period_cols count={len(period_cols)}")
                    for row in range(16, 56):
                        name = str(ws.cell(row=row, column=6).value or "").strip()
                        if not name:
                            continue
                        for period, col in period_cols.items():
                            val = ws.cell(row=row, column=col).value
                            if val is not None and val != 0 and val != "":
                                print(f"    Row {row}, period {period}, col {col}, val={val!r}")
                files.append(f)
                wb.close()
                continue
            ws = wb["Resources"]
            has_future = False
            # Read period headers from row 15
            period_cols = {}
            for col in range(10, 60):
                cell = ws.cell(row=15, column=col)
                period = parse_period_label(cell.value)
                if period and period >= FROM_PERIOD:
                    period_cols[period] = col
            # Check if any cell in the allocation area has a value >= FROM_PERIOD
            for row in range(16, 56):
                name = str(ws.cell(row=row, column=6).value or "").strip()
                if not name:
                    continue
                for period, col in period_cols.items():
                    val = ws.cell(row=row, column=col).value
                    try:
                        if val and float(val) > 0:
                            has_future = True
                            break
                    except (ValueError, TypeError):
                        pass
                if has_future:
                    break
            wb.close()
            if has_future:
                files.append(f)
            else:
                no_future.append(f.name)
            if not has_future or len(period_cols) == 0:
                print(f"  CHECK: {f.name} — period_cols={sorted(period_cols.keys())[:3]}, has_future={has_future}")
                print(f"  EXCLUDED: {f.name}")
        except Exception as e:
            print(f"  WARNING: Could not read {f.name}: {e}")
            no_future.append(f.name)

    print(f"  {len(files)} files have future allocations")
    print(f"  {len(no_future)} files excluded (no future work)\n")

    print(f"\nCTC Migration Tool")
    print(f"{'='*60}")
    print(f"  Folder:      {CTC_FOLDER}")
    print(f"  Database:    {DB_PATH}")
    print(f"  From period: {FROM_PERIOD}")
    print(f"  Files found: {len(files)}")
    print(f"  Mode:        {'DRY RUN — no changes will be made' if DRY_RUN else 'LIVE — writing to database'}")
    print(f"{'='*60}\n")

    if not DRY_RUN:
        confirm = input("Type YES to proceed with live import: ").strip()
        if confirm != "YES":
            print("Aborted.")
            sys.exit(0)

    if not preflight_check(files):
        confirm = input("Duplicates found. Type YES to continue anyway, or press Enter to abort: ").strip()
        if confirm != "YES":
            print("Aborted.")
            sys.exit(0)

    conn = get_conn() if not DRY_RUN else None

    stats = {
        "ok":           0,
        "skipped":      0,
        "errors":       0,
        "staff_skipped": [],
    }

    results = []
    for i, f in enumerate(files, 1):
        print(f"[{i:3d}/{len(files)}] {f.name[:50]:<50}", end=" ")
        result = process_file(f, conn, stats)
        status = result.get("status", "?")
        if status == "OK":
            stats["ok"] += 1
            print(f"OK   — {result.get('project_name','?')} "
                  f"({result.get('staff',0)} staff, starts {result.get('start_period','?')}, {result.get('rows_added',0)} allocation rows)")
        elif status == "DRY_RUN":
            stats["ok"] += 1
            print(f"PREVIEW — {result.get('proj_num','?')}/{result.get('task_num','?')} "
                  f"({result.get('staff',0)} staff, starts {result.get('start_period','?')}, {result.get('grid_months',12)}-month grid)")
        elif status == "COLLISION":
            stats["collisions"] = stats.get("collisions", 0) + 1
            print(f"COLLISION — {result.get('error','')}")
        elif status == "SKIPPED":
            stats["skipped"] += 1
            print(f"SKIP — {result.get('error','')}")
        else:
            stats["errors"] += 1
            print(f"ERROR — {result.get('error','')}")
        results.append(result)

    if conn:
        conn.close()

    print(f"\n{'='*60}")
    print(f"  Processed:     {stats['ok']}")
    print(f"  Collisions:    {stats.get('collisions', 0)}")
    print(f"  Skipped:       {stats['skipped']}")
    print(f"  Errors:        {stats['errors']}")
    if stats["staff_skipped"]:
        unique_skipped = sorted(set(stats["staff_skipped"]))
        print(f"  Staff not found in database ({len(unique_skipped)}):")
        for name in unique_skipped[:20]:
            print(f"    - {name}")
        if len(unique_skipped) > 20:
            print(f"    ... and {len(unique_skipped)-20} more")
    print(f"{'='*60}\n")

    if DRY_RUN:
        print("This was a DRY RUN. Set DRY_RUN = False to write to the database.")


if __name__ == "__main__":
    # Add the RFT app directory to the path so we can import database.py
    rft_dir = str(Path(DB_PATH).parent.parent)
    if rft_dir not in sys.path:
        sys.path.insert(0, rft_dir)
    main()
